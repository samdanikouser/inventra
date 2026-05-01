#!/usr/bin/env python3
"""
import_book2.py — Import items from Book2.xlsx into Inventra.

Usage:
    python3 import_book2.py --api-url http://localhost:8000/api \
                            --username admin --password yourpassword

The script will:
  1. Read Book2.xlsx (SKU + name columns)
  2. Auto-categorize items based on name keywords
  3. Create missing categories via the API
  4. Create items via the API (auto-generate SKU if missing)
  5. Skip items that already exist (by SKU)
"""

import argparse
import re
import sys

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required.  pip3 install openpyxl")
    sys.exit(1)

try:
    import requests
except ImportError:
    print("ERROR: requests is required.  pip3 install requests")
    sys.exit(1)


# ── Category classification rules ──────────────────────────────────────────
# Order matters — first match wins.
CATEGORY_RULES = [
    # Glassware & Drinkware
    ("Glassware", [
        r"glass", r"tumbler", r"goblet", r"flute", r"martini", r"whisky",
        r"shot", r"hiball", r"highball", r"cognac", r"brandy", r"longdrink",
        r"champagne", r"wine", r"burgundy", r"gin tonic", r"rocks",
        r"cocktail.*clear", r"old fashioned",
    ]),
    # Bar Equipment & Accessories
    ("Bar Equipment", [
        r"shaker", r"muddler", r"strainer", r"mixing spoon", r"measuring",
        r"julep", r"bar set", r"ice.*mould", r"ice.*bucket", r"carafe",
        r"decanter", r"jigger", r"mixing jug", r"bar scoop", r"soda.*siphon",
        r"smoker hood",
    ]),
    # Cutlery & Utensils
    ("Cutlery", [
        r"spoon", r"fork", r"tong", r"knife", r"cutlery set", r"ladle",
        r"serving tong", r"solid spoon", r"cake tong", r"chopstick",
    ]),
    # Plates (Flat, Deep, Oval, etc.)
    ("Plates", [
        r"flat plate", r"deep plate", r"oval plate", r"oval dish",
        r"pasta plate", r"dinner.*plate", r"saucer", r"rectangular plate",
        r"walled plate", r"oblong plate", r"gourmet.*plate", r"bloom.*plate",
        r"vago.*plate", r"hygge.*plate", r"banquet.*plate",
    ]),
    # Bowls
    ("Bowls", [
        r"bowl", r"ramen.*bowl", r"soup.*bowl", r"salad.*bowl",
        r"dipping.*bowl", r"tapas.*dish",
    ]),
    # Cups & Mugs
    ("Cups & Mugs", [
        r"cup", r"mug", r"coffee.*cup", r"tea.*server", r"teapot",
        r"teabloom", r"bodum", r"kettle", r"matcha.*bowl",
    ]),
    # Serving & Presentation
    ("Serving & Presentation", [
        r"server", r"stand", r"riser", r"platter", r"tray", r"boat",
        r"carrier", r"display.*dome", r"food cover", r"buffet",
        r"tier.*stand", r"cake.*stand", r"pastries.*stand", r"geta",
        r"sushi.*geta", r"menu.*holder", r"card.*holder",
    ]),
    # Wooden Serveware
    ("Wooden Serveware", [
        r"wood", r"walnut", r"bamboo", r"oak", r"hinoki", r"lacquer",
        r"chirashi.*box", r"bento.*box", r"sake.*box", r"temaki",
        r"edobitsu", r"sushi.*oke",
    ]),
    # Cast Iron & Hot Plates
    ("Cast Iron & Hot Plates", [
        r"cast iron", r"hot plate", r"lodge", r"lava", r"skillet",
        r"fajita", r"konro", r"grill", r"mortar.*pestle", r"granite.*mortar",
        r"molcajete",
    ]),
    # Condiment & Table Accessories
    ("Table Accessories", [
        r"salt.*shaker", r"pepper.*shaker", r"pepper.*mill", r"salt.*mill",
        r"chip.*pot", r"gravy.*boat", r"creamer", r"sauce.*pot",
        r"pail", r"moscow mule", r"pineapple cup",
    ]),
    # Steamer & Asian Cookware
    ("Asian Cookware", [
        r"steamer", r"tortilla", r"taco", r"wok",
    ]),
    # Pitchers, Carafes & Dispensers
    ("Pitchers & Dispensers", [
        r"pitcher", r"dispenser", r"jar",
    ]),
    # Slate & Stone
    ("Slate & Stone", [
        r"slate", r"stone", r"lava stone",
    ]),
    # Table Décor & Lighting
    ("Table Décor", [
        r"table lamp", r"placemat", r"tissue.*box", r"ashtray", r"ash.*tray",
        r"bread.*bag",
    ]),
    # Miscellaneous
    ("Miscellaneous Equipment", [
        r"scaler", r"net.*replacement", r"prep.*geta", r"neta.*case",
    ]),
]


def classify_item(name: str) -> str:
    """Return the best category name for a given item name."""
    lower = name.lower()
    for category, patterns in CATEGORY_RULES:
        for pat in patterns:
            if re.search(pat, lower):
                return category
    return "Uncategorised"


def generate_sku(name: str, index: int) -> str:
    """Generate a placeholder SKU from the item name."""
    # Take first letters of each word, uppercase, max 6 chars
    words = re.findall(r'[A-Za-z]+', name)
    prefix = "".join(w[0] for w in words[:4]).upper()
    if not prefix:
        prefix = "ITEM"
    return f"TMP-{prefix}-{index:04d}"


def main():
    parser = argparse.ArgumentParser(description="Import Book2.xlsx into Inventra")
    parser.add_argument("--api-url", required=True, help="e.g. http://localhost:8000/api")
    parser.add_argument("--username", required=True)
    parser.add_argument("--password", required=True)
    parser.add_argument("--xlsx", default="Book2.xlsx", help="Path to Excel file")
    parser.add_argument("--dry-run", action="store_true", help="Preview without creating")
    args = parser.parse_args()

    base = args.api_url.rstrip("/")
    session = requests.Session()

    # ── 1. Read Excel ──────────────────────────────────────────────────────
    print(f"📄 Reading {args.xlsx}...")
    wb = openpyxl.load_workbook(args.xlsx)
    ws = wb.active
    raw_items = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        sku_raw = str(row[0]).strip() if row[0] else ""
        name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        if sku_raw == "None":
            sku_raw = ""
        if not name:
            continue
        # Clean up multiline names (some cells have line breaks)
        name = re.sub(r'\s+', ' ', name).strip()
        raw_items.append({"sku": sku_raw, "name": name})

    print(f"   Found {len(raw_items)} items ({sum(1 for i in raw_items if i['sku'])} with SKU, "
          f"{sum(1 for i in raw_items if not i['sku'])} without)\n")

    # ── 3. Classify items ──────────────────────────────────────────────────
    print("🏷️  Classifying items into categories...")
    category_items: dict[str, list] = {}
    for idx, item in enumerate(raw_items, start=1):
        cat = classify_item(item["name"])
        if not item["sku"]:
            item["sku"] = generate_sku(item["name"], idx)
            item["auto_sku"] = True
        else:
            item["auto_sku"] = False
        item["category"] = cat
        category_items.setdefault(cat, [])
        category_items[cat].append(item)

    for cat, items in sorted(category_items.items()):
        print(f"   {cat}: {len(items)} items")
    print()

    if args.dry_run:
        print("🔍 DRY RUN — printing first 5 items per category:\n")
        for cat, items in sorted(category_items.items()):
            print(f"  ── {cat} ──")
            for it in items[:5]:
                flag = " (auto-SKU)" if it["auto_sku"] else ""
                print(f"    [{it['sku']}] {it['name']}{flag}")
            if len(items) > 5:
                print(f"    ... and {len(items) - 5} more")
            print()
        print("Re-run without --dry-run to create items.")
        return

    # ── 3. Authenticate ────────────────────────────────────────────────────
    print("🔑 Authenticating...")
    resp = session.post(f"{base}/token/", json={
        "username": args.username,
        "password": args.password,
    })
    if resp.status_code != 200:
        print(f"   ❌ Login failed: {resp.status_code} {resp.text}")
        sys.exit(1)
    token = resp.json()["access"]
    session.headers["Authorization"] = f"Bearer {token}"
    print("   ✅ Logged in.\n")

    # ── 4. Fetch existing categories ───────────────────────────────────────
    print("📥 Fetching existing categories...")
    existing_cats: dict[str, int] = {}
    url = f"{base}/categories/"
    while url:
        resp = session.get(url)
        data = resp.json()
        for c in data.get("results", data if isinstance(data, list) else []):
            existing_cats[c["name"]] = c["id"]
        url = data.get("next")
    print(f"   Found {len(existing_cats)} existing: {', '.join(existing_cats.keys()) or '(none)'}\n")

    # ── 5. Create missing categories ───────────────────────────────────────
    needed_cats = set(category_items.keys()) - set(existing_cats.keys())
    if needed_cats:
        print(f"➕ Creating {len(needed_cats)} new categories...")
        for cat_name in sorted(needed_cats):
            resp = session.post(f"{base}/categories/", json={"name": cat_name})
            if resp.status_code == 201:
                existing_cats[cat_name] = resp.json()["id"]
                print(f"   ✅ {cat_name}")
            else:
                print(f"   ⚠️  {cat_name}: {resp.status_code} {resp.text}")
        print()

    # ── 6. Fetch existing items (to skip duplicates) ───────────────────────
    print("📥 Fetching existing items...")
    existing_skus: set[str] = set()
    url = f"{base}/items/"
    while url:
        resp = session.get(url)
        data = resp.json()
        for it in data.get("results", data if isinstance(data, list) else []):
            existing_skus.add(it["sku"])
        url = data.get("next")
    print(f"   Found {len(existing_skus)} existing items.\n")

    # ── 7. Create items ────────────────────────────────────────────────────
    created = 0
    skipped = 0
    failed = 0

    print("📦 Creating items...")
    for cat_name, items in sorted(category_items.items()):
        cat_id = existing_cats.get(cat_name)
        if not cat_id:
            print(f"   ⚠️  No category ID for '{cat_name}' — skipping {len(items)} items")
            failed += len(items)
            continue

        for item in items:
            if item["sku"] in existing_skus:
                skipped += 1
                continue

            payload = {
                "sku": item["sku"],
                "name": item["name"],
                "category": cat_id,
                "unit": "pcs",
                "min_stock": 5,
                "par_level": 20,
                "unit_cost": "0.000",
            }
            resp = session.post(f"{base}/items/", json=payload)
            if resp.status_code == 201:
                created += 1
                existing_skus.add(item["sku"])
                flag = " 🔖 (auto-SKU, editable)" if item["auto_sku"] else ""
                if created <= 20 or created % 50 == 0:
                    print(f"   ✅ [{item['sku']}] {item['name'][:50]}{flag}")
            else:
                failed += 1
                print(f"   ❌ [{item['sku']}] {resp.status_code}: {resp.text[:120]}")

    print(f"\n{'='*60}")
    print(f"📊 IMPORT COMPLETE")
    print(f"   ✅ Created:  {created}")
    print(f"   ⏭️  Skipped:  {skipped} (already exist)")
    print(f"   ❌ Failed:   {failed}")
    print(f"   📂 Categories: {len(existing_cats)}")
    print(f"{'='*60}")

    if any(it["auto_sku"] for items in category_items.values() for it in items):
        print(f"\n💡 NOTE: {sum(1 for items in category_items.values() for it in items if it['auto_sku'])} items "
              f"have auto-generated SKUs (prefix TMP-*).")
        print("   You can edit them in the Inventory page → click the pencil icon on each item.")


if __name__ == "__main__":
    main()
